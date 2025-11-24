from datetime import datetime
import os, json

from django.core.management.base import BaseCommand



class Command(BaseCommand):
    help = 'Generate cans client report'

    def handle(self, *args, **options):
        self.stdout.write(f'Generating items')

        try:
            file_neame = create_items_excel_from_django_data()
        except Exception as e: 
            self.stdout.write(e)

        self.stdout.write(
            self.style.SUCCESS(file_neame)
        )









from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from items.models import Items

def create_items_excel_from_django_data(data = Items.objects.all(), filename="items_export.xlsx"):
    """
    Create Excel file from Django model data with dynamic repository columns
    
    Args:
        data: List of item dictionaries from Django serializer
        filename: Output Excel filename
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Export"
    
    # Define styles
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    subheader_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_font = Font(color="000000", size=10)
    data_font = Font(color="000000", size=9)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all unique repositories from the data
    all_repositories = set()
    for item in data:
        for stock in item.stock.all():
            all_repositories.add(stock.repository.name)
    
    all_repositories = sorted(list(all_repositories))  # Sort for consistent order
    
    # Define base columns
    base_columns = [
        "Item ID", "Name", "Type", "By Username", "Created At", "Updated At",
        "Price 1", "Price 2", "Price 3", "Price 4", "Origin", "Place", 
        "Is Refillable", "Barcodes", "Images"
    ]
    
    # Create dynamic headers
    headers = base_columns.copy()
    
    # Add repository columns
    for repo in all_repositories:
        headers.append(f"Stock - {repo}")
    
    # Write main headers (row 1)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Create subheaders (row 2) - more descriptive
    subheaders = [
        "ID", "Item Name", "Type Name", "User", "Creation Date", "Last Update",
        "Price 1", "Price 2", "Price 3", "Price 4", "Origin", "Location", 
        "Refillable", "Barcode List", "Image URLs"
    ]
    
    # Add quantity subheaders for repositories
    for repo in all_repositories:
        subheaders.append("Quantity")
    
    for col, subheader in enumerate(subheaders, 1):
        cell = ws.cell(row=2, column=col, value=subheader)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data starting from row 3
    for row_idx, item in enumerate(data, 3):
        col = 1
        
        # Basic item information
        basic_data = [
            str(item.id) if item.id else "",
            str(item.name) if item.name else "",
            str(item.type_name) if hasattr(item, 'type_name') and item.type_name else (str(item.type.name) if hasattr(item, 'type') and item.type else ''),
            str(item.by_username) if hasattr(item, 'by_username') and item.by_username else (str(item.by.username) if hasattr(item, 'by') and item.by else ''),
            format_datetime(item.created_at),
            format_datetime(item.updated_at),
            str(item.price1) if item.price1 else "0.00",
            str(item.price2) if item.price2 else "0.00", 
            str(item.price3) if item.price3 else "0.00",
            str(item.price4) if item.price4 else "0.00",
            str(item.origin) if item.origin else "",
            str(item.place) if item.place else "",
            'Yes' if item.is_refillable else 'No',
            format_barcodes_django(item.barcodes.all()),
            format_images_django(getattr(item, 'images', None))
        ]
        
        # Write basic data
        for value in basic_data:
            # Convert any object to string for Excel compatibility
            excel_value = str(value) if value is not None else ""
            cell = ws.cell(row=row_idx, column=col, value=excel_value)
            cell.font = data_font
            cell.border = border
            col += 1
        
        # Create stock dictionary for easy lookup
        stock_dict = {}
        for stock in item.stock.all():
            stock_dict[stock.repository.name] = stock.quantity
        
        # Write stock quantities for each repository
        for repo in all_repositories:
            quantity = stock_dict.get(repo, '0.00')
            cell = ws.cell(row=row_idx, column=col, value=str(quantity))
            cell.font = data_font
            cell.border = border
            col += 1
    
    # Auto-adjust column widths
    adjust_column_widths(ws)
    
    path = f'media/items_as_xlsx/{datetime.now()}-{filename}'
    # Save workbook
    wb.save(path)
    # print(f"Excel file created successfully: {path}")
    return path

def format_datetime(datetime_obj):
    """Format Django datetime object to readable format"""
    if not datetime_obj:
        return ""
    try:
        return datetime_obj.strftime('%Y-%m-%d %H:%M:%S')
    except:
        return str(datetime_obj)

def format_barcodes_django(barcodes_queryset):
    """Format Django barcodes queryset to string"""
    if not barcodes_queryset.exists():
        return ""
    return ", ".join([barcode.barcode for barcode in barcodes_queryset])

def format_images_django(images_relation):
    """Format Django images relation to string"""
    if not images_relation:
        return ""
    
    try:
        # If it's a related manager, get all objects
        if hasattr(images_relation, 'all'):
            images_list = list(images_relation.all())
        elif hasattr(images_relation, '__iter__'):
            images_list = list(images_relation)
        else:
            return ""
            
        if len(images_list) == 0:
            return ""
        elif len(images_list) == 1:
            # Try to get URL or string representation
            image = images_list[0]
            if hasattr(image, 'url'):
                return str(image.url)
            elif hasattr(image, 'image') and hasattr(image.image, 'url'):
                return str(image.image.url)
            else:
                return str(image)
        else:
            # Multiple images - show count and first image
            first_image = images_list[0]
            if hasattr(first_image, 'url'):
                first_url = str(first_image.url)
            elif hasattr(first_image, 'image') and hasattr(first_image.image, 'url'):
                first_url = str(first_image.image.url)
            else:
                first_url = str(first_image)
            return f"{len(images_list)} images: {first_url}"
    except Exception as e:
        return f"Error loading images: {str(e)}"

def adjust_column_widths(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set width with some padding, max 50 characters
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_advanced_items_excel(data, filename="advanced_items_export.xlsx"):
    """
    Advanced version with separate sheets for different data aspects
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create main items sheet
    create_main_items_sheet(wb, data)
    
    # Create stock details sheet
    create_stock_details_sheet(wb, data)
    
    # Create images sheet
    create_images_sheet(wb, data)
    
    # Create barcodes sheet
    create_barcodes_sheet(wb, data)
    
    wb.save(filename)
    print(f"Advanced Excel file created: {filename}")
    return filename

def create_main_items_sheet(wb, data):
    """Create main items information sheet"""
    ws = wb.create_sheet("Items")
    
    # Define styles (same as above)
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = [
        "Item ID", "Name", "Type", "By Username", "Created At", "Updated At",
        "Price 1", "Price 2", "Price 3", "Price 4", "Origin", "Place", "Is Refillable"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Write data
    for row_idx, item in enumerate(data, 2):
        data_row = [
            item.get('id', ''),
            item.get('name', ''),
            item.get('type_name', ''),
            item.get('by_username', ''),
            format_datetime(item.get('created_at', '')),
            format_datetime(item.get('updated_at', '')),
            item.get('price1', '0.00'),
            item.get('price2', '0.00'),
            item.get('price3', '0.00'),
            item.get('price4', '0.00'),
            item.get('origin', ''),
            item.get('place', ''),
            'Yes' if item.get('is_refillable', False) else 'No'
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
    
    adjust_column_widths(ws)

def create_stock_details_sheet(wb, data):
    """Create detailed stock information sheet"""
    ws = wb.create_sheet("Stock Details")
    
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Item ID", "Item Name", "Repository Name", "Quantity", "Repository ID"]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Write stock data
    row_idx = 2
    for item in data:
        for stock in item.get('stock', []):
            data_row = [
                item.get('id', ''),
                item.get('name', ''),
                stock.get('repository_name', ''),
                stock.get('quantity', ''),
                stock.get('repository', '')
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
            row_idx += 1
    
    adjust_column_widths(ws)

def create_images_sheet(wb, data):
    """Create images sheet"""
    ws = wb.create_sheet("Images")
    
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Item ID", "Item Name", "Image URL"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row_idx = 2
    for item in data:
        for image_url in item.get('images', []):
            data_row = [
                item.get('id', ''),
                item.get('name', ''),
                image_url
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
            row_idx += 1
    
    adjust_column_widths(ws)

def create_barcodes_sheet(wb, data):
    """Create barcodes sheet"""
    ws = wb.create_sheet("Barcodes")
    
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Item ID", "Item Name", "Barcode ID", "Barcode"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row_idx = 2
    for item in data:
        for barcode in item.get('barcodes', []):
            data_row = [
                item.get('id', ''),
                item.get('name', ''),
                barcode.get('id', ''),
                barcode.get('barcode', '')
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
            row_idx += 1
    
    adjust_column_widths(ws)

# Example usage
if __name__ == "__main__":
    # Sample data based on your Django model
    sample_data = [
        {
            "id": 1000000,
            "by_username": "استنفار",
            "images": [
                "http://localhost:8000/media/items-images/03a27493-ab2.png",
                "http://localhost:8000/media/items-images/f85e2ecc-d21.png",
                "http://localhost:8000/media/items-images/61c98874-a63.jpg"
            ],
            "stock": [
                {
                    "id": 1,
                    "repository_name": "الرئيسي",
                    "quantity": "8.00",
                    "item": 1000000,
                    "repository": 10000
                }
            ],
            "barcodes": [
                {
                    "id": 1,
                    "barcode": "7122-00-9552"
                }
            ],
            "type_name": "cans",
            "created_at": "2025-03-13T08:57:32.718082Z",
            "updated_at": "2025-06-19T09:52:37.753370Z",
            "name": "12/15/18g LENS cap pro ##USA, Turkiye @@9",
            "price1": "0.00",
            "price2": "0.00",
            "price3": "0.00",
            "price4": "0.00",
            "is_refillable": False,
            "origin": "ewf",
            "place": "fds",
            "by": 10000,
            "type": 1
        }
    ]
    
    # Create simple Excel
    create_items_excel_from_django_data(sample_data, "items_simple.xlsx")
    
    # Create advanced Excel with multiple sheets
    create_advanced_items_excel(sample_data, "items_advanced.xlsx")
    
    print("Excel files created successfully!")
