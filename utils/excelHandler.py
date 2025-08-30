import openpyxl as xl

def proccess_workbook(filename = 'initial_data.xlsx', sheetname = 'Sheet1'):
    return
    wb = xl.load_workbook(filename)
    sheet = wb[sheetname]

    for row in range(4, sheet.max_row + 2):
        item = {}
        for col in range(3, sheet.max_column + 1):
            # create key value bears
            key = sheet.cell(3, col).value# .lower().strip().replace(' ', '_')
            value = str(sheet.cell(row, col).value)#.lower().strip()

            item[key] = value
        
        if item['اسم الصنف'] != 'None':
            # append rows
            origin = '##' + str(item['بلد المنشأ']) if item['بلد المنشأ'] != 'None' else ''
            i = Items.objects.create(
                name=item['اسم الصنف'] + f' {origin}',
                by_id=10000
            )
            if item['كود الصنف'] != 'None':
                i.barcodes.create(
                    barcode=item['كود الصنف']
                )
            if item['العدد'] != 'None':
                i.initial_stock.create(
                    repository_id=10000,
                    quantity=item['العدد'],
                )


from invoices.buyer_supplier_party.models import Party, InitialCreditBalance
from refillable_items_system.models import RefillableItemsInitialStockClientHas, RefundedRefillableItem, ItemTransformer
from finance.payments.models import Payment
from invoices.sales.models import SalesInvoice, SalesInvoiceItem
# from invoices.purchase.models import PurchaseInvoices
from datetime import timedelta
from django.db import transaction
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP


def proccess_workbook_clients(filename = 'عملاء 2025.xlsx'):
    with transaction.atomic():
        wb = xl.load_workbook(filename)
        for sheet in wb.worksheets:
            print(11111111111)
            # print(sheet.title)
            # print(1000+sheet.cell(1, 7).value)
            

            print(sheet.cell(1, 7).value)
            item = ItemTransformer.objects.get(transform__id=int(1000000)+int(sheet.cell(1, 7).value)).item
            owner = Party.objects.create(
                name=sheet['A1'].value,
                detail='',
                created_at=timezone.now(),
                by_id=10000,
            )
            if sheet['S2'].value > 0:
                InitialCreditBalance.objects.create(
                    party=owner,
                    amount=sheet['S2'].value,
                    date=timezone.now(),
                    created_at=timezone.now(),
                    by_id=10000,
                )
            if sheet['S4'].value > 0:
                RefillableItemsInitialStockClientHas.objects.create(
                    item=item,
                    owner=owner,
                    quantity=sheet['S4'].value,
                    date=timezone.now(),
                    created_at=timezone.now(),
                    by_id=10000,
                )
            a = None
            for row in range(6, sheet.max_row + 2):
                if not sheet.cell(row, 2).value:
                    break
                if sheet.cell(row, 3).value:
                    
                    if sheet.cell(row, 2).value != a:
                        print(sheet.title)
                        print(sheet.cell(row, 3).value)
                        print(sheet.cell(row, 5).value)
                        a = sheet.cell(row, 2).value
                        g = sheet.cell(row, 3).value * int(sheet.cell(row, 5).value.split('*')[1])
                        v = Decimal(str(g)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        # print(v)
                        inv = SalesInvoice.objects.create(
                            owner=owner,
                            issue_date=sheet.cell(row, 2).value,
                            due_date=sheet.cell(row, 2).value + timedelta(days=14),
                            total_amount=v,
                            repository_permit=True,
                            created_at=timezone.now(),
                            by_id=10000,
                        )
                    print(22222222222222)
                    SalesInvoiceItem.objects.create(
                        invoice=inv,
                        repository_id=10000,
                        item_id=int(1000000)+int(sheet.cell(1, 7).value),
                        
                        quantity=sheet.cell(row, 3).value,
                        unit_price=int(sheet.cell(row, 5).value.split('*')[1]),
                        # total=int(sheet.cell(row, 5).value) * int(sheet.cell(row, 3).value),
                        # discount=sheet.cell(row, 5).value,
                        # tax_rate=sheet.cell(row, 6).value,
                    )
                    print(4444444444444)
                print(3333333333)
                if sheet.cell(row, 4).value:
                    RefundedRefillableItem.objects.create(
                        item=item,
                        owner=owner,
                        repository_id=10000,
                        quantity=sheet.cell(row, 4).value,
                        date=sheet.cell(row, 2).value,
                        created_at=timezone.now(),
                        by_id=10000,
                    )
                print(5555555555555)    
            for rowww in range(10, sheet.max_row + 2):
                if not sheet.cell(rowww, 18).value:
                    break
                Payment.objects.create(
                    owner=owner,
                    amount=sheet.cell(rowww, 18).value,
                    payment_method_id=1000000,
                    date=sheet.cell(rowww, 17).value,
                    paid=True,
                    created_at=timezone.now(),
                    by_id=10000,
                )
            print(66666666666)

            

# proccess_workbook_clients()
